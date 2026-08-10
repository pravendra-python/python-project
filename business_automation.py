import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Business data
data = {
    "Customer": ["Rahul", "Amit", "Neha", "Pooja"],
    "Product": ["Laptop", "Mouse", "Keyboard", "Laptop"],
    "Quantity": [1, 3, 2, 2],
    "Price": [50000, 500, 1200, 50000]
}

# Create DataFrame
df = pd.DataFrame(data)

# Calculate total
df["Total"] = df["Quantity"] * df["Price"]

# Business calculations
total_revenue = df["Total"].sum()

product_sales = df.groupby("Product")["Quantity"].sum()
best_selling_product = product_sales.idxmax()

# Create Excel report
file_name = "business_sales_report.xlsx"

with pd.ExcelWriter(file_name, engine="openpyxl") as writer:

    df.to_excel(writer, sheet_name="Sales Data", index=False)

    summary = pd.DataFrame({
        "Metric": [
            "Total Revenue",
            "Total Quantity Sold",
            "Best-Selling Product"
        ],
        "Value": [
            total_revenue,
            df["Quantity"].sum(),
            best_selling_product
        ]
    })

    summary.to_excel(writer, sheet_name="Business Summary", index=False)

# Open Excel file for formatting
workbook = load_workbook(file_name)

# Format both sheets
for sheet in workbook.worksheets:

    # Header
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )
        cell.alignment = Alignment(horizontal="center")

    # Column width
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[column_letter].width = max_length + 3

# Save final report
workbook.save(file_name)

print("Professional Excel report created successfully!")
print("File name:", file_name)
print("Total Revenue:", total_revenue)
print("Best-Selling Product:", best_selling_product)
